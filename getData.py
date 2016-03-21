import os
import openpyxl
import sys
from bs4 import BeautifulSoup


def ParseDataFromHtml(html):
	soup = BeautifulSoup(html,'lxml')
	table = soup.find('table',{'id':'gvSearchResult'})
	rows = table.findChildren('tr')
	result= list()
	for row in rows:
		cells = row.findChildren('td')
		i=0
		for cell in cells:
	
			if(i<5):      #first 5 cells has useless data, hence skip them
				i+=1
				continue

			value = cell.string
			result.append(value)
			i+=1

	return result

def getData(file):
	command1 = "curl 'http://164.100.180.4/searchengine/SearchEngineEnglish.aspx' -H 'Cookie: ASP.NET_SessionId=wvtex4icq3ez20yniaw31555' -H 'Origin: http://164.100.180.4' -H 'Accept-Encoding: gzip, deflate' -H 'Accept-Language: en,af;q=0.8,en-US;q=0.6' -H 'Upgrade-Insecure-Requests: 1' -H 'User-Agent: Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.87 Safari/537.36' -H 'Content-Type: application/x-www-form-urlencoded' -H 'Accept: text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8' -H 'Cache-Control: max-age=0' -H 'Referer: http://164.100.180.4/searchengine/SearchEngineEnglish.aspx' -H 'Connection: keep-alive' --data '__EVENTTARGET=&__EVENTARGUMENT=&__LASTFOCUS=&__VIEWSTATE=%2FwEPDwULLTExMjYxOTk3OTEPZBYCAgEPZBYQAgEPEGRkFgFmZAIDDxAPFgYeDkRhdGFWYWx1ZUZpZWxkBQtEaXN0cmljdF9JRB4NRGF0YVRleHRGaWVsZAUQRGlzdHJpY3RfTmFtZV9Fbh4LXyFEYXRhQm91bmRnZBAVTAotLVNlbGVjdC0tBEFncmEHQWxpZ2FyaAlBbGxhaGFiYWQOQW1iZWRrYXIgTmFnYXIGQW1ldGhpBkFtcm9oYQdBdXJhaXlhCEF6YW1nYXJoB0JhZ2hwYXQJQmFoYXJhaWNoBUJhbGlhCUJhbHJhbXB1cgVCYW5kYQlCYXJhYmFua2kIQmFyZWlsbHkFQmFzdGkGQmlqbm9yBkJ1ZGF1bgtCdWxhbmRzYWhhcglDaGFuZGF1bGkKQ2hpdHJha29vdAZEZW9yaWEERXRhaAZFdGF3YWgIRmFpemFiYWQLRmFycnVraGFiYWQIRmF0ZWhwdXIJRmlyb3phYmFkE0dhdXRhbSBCdWRkaGEgTmFnYXIJR2hhemlhYmFkCEdoYXppcHVyBUdvbmRhCUdvcmFraHB1cghIYW1pcnB1cgVIYXB1cgZIYXJkb2kHSGF0aHJhcwZKYWxhdW4HSmF1bnB1cgZKaGFuc2kHS2FubmF1agxLYW5wdXIgRGVoYXQMS2FucHVyIE5hZ2FyB0thc2dhbmoJS2F1c2hhbWJpBUtoZXJpCkt1c2hpbmFnYXIITGFsaXRwdXIHTHVja25vdwtNYWhhcmFqZ2FuagZNYWhvYmEHTWFucHVyaQdNYXRodXJhA01hdQZNZWVydXQITWlyemFwdXIJTW9yYWRhYmFkDU11emFmZmFybmFnYXIIUGlsaWJoaXQKUHJhdGFwZ2FyaApSYWUgQmFyZWxpBlJhbXB1cgpTYWhhcmFucHVyB1NhbWJoYWwQU2FudCBLYWJpciBOYWdhchJTYW50IFJhdmlkYXMgTmFnYXIMU2hhaGphaGFucHVyBlNoYW1saQlTaHJhd2FzdGkOU2lkZGhhcnRobmFnYXIHU2l0YXB1cglTb25iaGFkcmEJU3VsdGFucHVyBVVubmFvCFZhcmFuYXNpFUwKLS1TZWxlY3QtLQIwOAIwOQIyOQI3MAI3MgIyMgIyOAI0NwIwNwI2NAI0OQI2NQIzOAI2OQIxNQI1NAIyMQIxNgIwNQI0MwIzOQI1MgIxMwIyNQI2NwIyNgIzMAIxMQIwNgIwNAI0MgI2MwI1MAIzNgI3MwI2MQIxNAIzNQI0MQIzMwIyNwIyNAIyMwI3MQIzMgI2MgI1MwIzNAI1NwI1MQIzNwIxMgIxMAI0OAIwMwI0NAIxOQIwMgIxOAIzMQI1OQIyMAIwMQI3NQI1NgI0NgIxNwI3NAI2NgI1NQI2MAI0NQI2OAI1OAI0MBQrA0xnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnZ2dnFgECGWQCBQ8WAh4HVmlzaWJsZWgWAgICD2QWAgIBDxBkZBYAZAIHDxYCHwNoFgICAg9kFgICAQ8PFgIeBFRleHRlZGQCCQ8WAh8DZxYCAgEPZBYOAgEPEGRkFgFmZAILDw8WAh8DaGRkAg0PDxYEHwRlHwNoZGQCDw8PFgIfA2hkZAIRDxAPFgIfA2hkZBYBZmQCEw8PFgIfA2hkZAIVDw8WBB8EZR8DaGRkAg8PDxYCHwRlZGQCEQ88KwANAQAPFgIfA2hkZAITDzwrAA0BAA8WAh8DaGRkGAMFCFNlYXJjaGdkD2dkBQ5ndlNlYXJjaFJlc3VsdA9nZAUGY2NKb2luDwUkODkxYmJhZmYtMjEzMC00ZjU0LTg3Y2ItNTg5Nzk4MTQ3NTgzZMRoxiNvjS9u9LXUXyajnIbUq8cy&__VIEWSTATEGENERATOR=6C38677B&__EVENTVALIDATION=%2FwEWWQLZ64S0BQKC8MqlAwKd8MqlAwKc8MqlAwKSn%2BDLDwKHlN6LCAKl2dH%2FDQKX%2B9TlBAKX%2B9jlBAKJ%2B9jlBAKO%2B7TmBAKO%2B7zmBAKJ%2B7zmBAKJ%2B9TlBAKL%2B5DmBAKX%2B5DmBAKN%2B4TmBAKL%2B9jlBAKN%2B4jmBAKK%2B9TlBAKN%2B9jlBAKI%2B4jmBAKM%2B4TmBAKJ%2B7jmBAKI%2B4zmBAKX%2B4jmBAKL%2B4DmBAKK%2B9jlBAKM%2B7zmBAKI%2B4DmBAKJ%2B4jmBAKN%2B5DmBAKJ%2B4zmBAKK%2B7TmBAKI%2B7jmBAKX%2B4zmBAKX%2B4TmBAKL%2B7zmBAKN%2B4DmBAKM%2B7TmBAKK%2B4zmBAKO%2B4DmBAKN%2B7jmBAKI%2B4TmBAKK%2B4jmBAKL%2B7jmBAKK%2B4DmBAKJ%2B5DmBAKJ%2B4TmBAKJ%2B4DmBAKO%2B7jmBAKK%2B7zmBAKN%2B7zmBAKM%2B4DmBAKK%2B4TmBAKM%2B5DmBAKM%2B7jmBAKK%2B5DmBAKI%2B7zmBAKI%2B7TmBAKL%2B9TlBAKX%2B4DmBAKL%2B4TmBAKI%2B9jlBAKX%2B7zmBAKI%2B9TlBAKK%2B7jmBAKM%2B9jlBAKJ%2B7TmBAKX%2B7jmBAKO%2B4jmBAKM%2B4zmBAKL%2B4zmBAKI%2B5DmBAKO%2B4TmBAKN%2B4zmBAKM%2B4jmBAKN%2B7TmBAKL%2B4jmBAKN%2B9TlBAKM%2B9TlBAKL%2B7TmBAL7r%2BtjAuSv62MC5a%2FrYwLrwMGNDAKaz5n5DALvmdGiDAKM54rGBsdNmIve7UJ2pLFDdX3ZILOEPWT4&RdlSearch=0&ddlDistricts=67&RdlSearchBy=0&txtEPICNo="
	command2="&Button1=Search' --compressed"
	wb = openpyxl.load_workbook(file)
	ws = wb.get_sheet_by_name('Sheet1')
	NA = 0
	for i in range(2,len(ws.columns[0]) + 1):
		EPICNo = ws['A'+str(i)].value
		html = os.popen(command1+EPICNo+command2).read()
		#html = process.read()
		#process.close()
		if 'No Match Found' in html:
			NA = NA + 1
			ws['B'+str(i)] = 'NA'
			ws['C'+str(i)] = 'NA'
			ws['D'+str(i)] = 'NA'
			ws['E'+str(i)] = 'NA'
			ws['F'+str(i)] = 'NA'
			ws['G'+str(i)] = 'NA'
		else:
			result = ParseDataFromHtml(html)
			ws['B'+str(i)] = result[0]
			ws['C'+str(i)] = result[1]
			ws['D'+str(i)] = result[2]
			ws['E'+str(i)] = result[3]
			ws['F'+str(i)] = result[4]
			ws['G'+str(i)] = result[5]
	print ("NA = ", NA)
	wb.save(file)


if __name__ == "__main__":
	file = sys.argv[1]
	getData(file)